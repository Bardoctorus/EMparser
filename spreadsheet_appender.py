from openpyxl import load_workbook, Workbook
import os.path
import datetime




def append_sheet(valueToAdd):
    dateToday = datetime.date.today()
    filename="EMlist.xlsx"
    if(os.path.exists(filename)):
        workbook = load_workbook(filename)
        backupDate = dateToday.strftime("%Y_%m_%d")
        backupName = backupDate + "_backup.xlsx"
        workbook.save(backupName)
    else:
        workbook = Workbook()

    sheet = workbook.active

    sheet.cell(column=1, row=sheet.max_row+1,value = valueToAdd)
    sheet.cell(column=2, row=sheet.max_row,value = dateToday)

    workbook.save(filename)